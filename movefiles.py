import shutil
import os
import openpyxl

currentDirectory = os.getcwd()
excelPath = currentDirectory + '\\Config.xlsx'

open_workbook_obj = openpyxl.load_workbook(excelPath)
sheet_obj = open_workbook_obj.active
source_folder = sheet_obj.cell(row = 2, column = 1)
destination_folder = sheet_obj.cell(row = 2, column = 2)

folder_names = ['exe','pdf','xlsx','zip','jpg']

fullpath = os.path.join

path_dict = {}

for directories in folder_names: 
    new_path = os.path.join(destination_folder.value,directories.upper())
    os.makedirs(new_path,exist_ok=True)
    path_dict[directories] = new_path


count = 0
for file in os.listdir(source_folder.value):
    try:
        ext = os.path.splitext(file)[1][1:]
        if ext in path_dict:
            src = (os.path.join(source_folder.value, file))
            dst = path_dict[ext]
            shutil.move(src,dst)
            count = count + 1
    except:
        print(f'File already exists {file}')
 
print(f'The number of files moved were : {count} ')



